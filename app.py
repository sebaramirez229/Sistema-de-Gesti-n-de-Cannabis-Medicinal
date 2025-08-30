import streamlit as st
import pandas as pd
import plotly.express as px
from datetime import datetime
import io
import os
from openpyxl import load_workbook, Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# Configuración de la página
st.set_page_config(
    page_title="Sistema de Gestión de Cannabis Medicinal",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Título principal
st.title("🌿 Sistema de Gestión de Cannabis Medicinal")
st.markdown("---")

# Configuración en el sidebar
st.sidebar.header("Configuración")

# Cargar archivos Excel personalizados
archivos_cargados = st.sidebar.file_uploader(
    "Cargar archivos Excel", 
    type=['xlsx'], 
    accept_multiple_files=True,
    help="Suba uno o más archivos Excel para crear nuevas vistas"
)

# Configuración de nombres de hojas
st.sidebar.subheader("Configuración de nombres de hojas")
nombres_hojas = {
    'Dispensarios': st.sidebar.text_input('Hoja de Dispensarios', value='Dispensarios'),
    'Alertas': st.sidebar.text_input('Hoja de Alertas', value='Alertas'),
    'Control_Calidad': st.sidebar.text_input('Hoja de Control de Calidad', value='Control_Calidad'),
    'Inventario_Deposito': st.sidebar.text_input('Hoja de Inventario Depósito', value='Inventario_Deposito'),
    'Inventario_Dispensario': st.sidebar.text_input('Hoja de Inventario Dispensario', value='Inventario_Dispensario'),
    'Ventas': st.sidebar.text_input('Hoja de Ventas', value='Ventas'),
    'Detalle_Venta': st.sidebar.text_input('Hoja de Detalle de Ventas', value='Detalle_Venta'),
    'Productos': st.sidebar.text_input('Hoja de Productos', value='Productos'),
    'Clientes': st.sidebar.text_input('Hoja de Clientes', value='Clientes')
}

# Configuración de colores
st.sidebar.subheader("Configuración de colores")
color_principal = st.sidebar.color_picker('Color principal', '#00CC96')
color_secundario = st.sidebar.color_picker('Color secundario', '#636EFA')

# Función para cargar datos
def load_data(archivos=None):
    datos = {}
    
    # Siempre cargar el archivo por defecto si existe
    try:
        if os.path.exists('data/db.xlsx'):
            excel_por_defecto = pd.read_excel('data/db.xlsx', sheet_name=None)
            for hoja, df in excel_por_defecto.items():
                datos[f"PorDefecto_{hoja}"] = df
    except Exception as e:
        st.sidebar.error(f"Error cargando archivo por defecto: {e}")
    
    # Si se cargaron archivos, usarlos
    if archivos:
        for archivo in archivos:
            try:
                excel_data = pd.read_excel(archivo, sheet_name=None)
                # Agregar prefijo al nombre de las hojas para identificar el archivo
                for hoja, df in excel_data.items():
                    nombre_archivo = archivo.name.replace('.xlsx', '')
                    datos[f"{nombre_archivo}_{hoja}"] = df
            except Exception as e:
                st.error(f"Error cargando {archivo.name}: {e}")
    return datos

# Función para guardar datos en el archivo Excel
def save_to_excel(df, sheet_name, filename='data/db.xlsx'):
    try:
        # Verificar si el directorio existe, si no, crearlo
        os.makedirs(os.path.dirname(filename), exist_ok=True)
        
        # Verificar si el archivo existe
        if os.path.exists(filename):
            # Cargar el libro existente
            book = load_workbook(filename)
            # Verificar si la hoja existe
            if sheet_name in book.sheetnames:
                # Eliminar la hoja existente
                std = book[sheet_name]
                book.remove(std)
            
            # Crear una nueva hoja con el DataFrame
            new_sheet = book.create_sheet(sheet_name)
            for r in dataframe_to_rows(df, index=False, header=True):
                new_sheet.append(r)
            
            # Guardar el libro
            book.save(filename)
        else:
            # Crear un nuevo archivo Excel
            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name=sheet_name, index=False)
        
        st.success(f"Datos guardados correctamente en {sheet_name}")
        return True
    except Exception as e:
        st.error(f"Error guardando datos: {e}")
        return False

# Cargar todos los datos
excel_data = load_data(archivos_cargados)

# Función para obtener datos por nombre de hoja
def obtener_datos(nombre_hoja):
    for clave, df in excel_data.items():
        # Buscar por el nombre de la hoja sin el prefijo del archivo
        if clave.endswith(f"_{nombre_hoja}") or clave == nombre_hoja:
            return df
    return None

# Obtener dataframes individuales
dispensarios_df = obtener_datos(nombres_hojas['Dispensarios'])
alertas_df = obtener_datos(nombres_hojas['Alertas'])
control_calidad_df = obtener_datos(nombres_hojas['Control_Calidad'])
inventario_deposito_df = obtener_datos(nombres_hojas['Inventario_Deposito'])
inventario_dispensario_df = obtener_datos(nombres_hojas['Inventario_Dispensario'])
ventas_df = obtener_datos(nombres_hojas['Ventas'])
detalle_venta_df = obtener_datos(nombres_hojas['Detalle_Venta'])
productos_df = obtener_datos(nombres_hojas['Productos'])
clientes_df = obtener_datos(nombres_hojas['Clientes'])

# Verificar que todos los datos necesarios estén disponibles
dataframes_requeridos = {
    'Dispensarios': dispensarios_df,
    'Alertas': alertas_df,
    'Control de Calidad': control_calidad_df,
    'Inventario Depósito': inventario_deposito_df,
    'Inventario Dispensario': inventario_dispensario_df,
    'Ventas': ventas_df,
    'Detalle de Ventas': detalle_venta_df,
    'Productos': productos_df,
    'Clientes': clientes_df
}

for nombre, df in dataframes_requeridos.items():
    if df is None:
        st.sidebar.warning(f"No se encontraron datos para: {nombre}")

# Sidebar con filtros
st.sidebar.header("Filtros")
if dispensarios_df is not None and 'nombre' in dispensarios_df.columns:
    dispensario_options = ["Todos"] + list(dispensarios_df['nombre'].unique())
    selected_dispensario = st.sidebar.selectbox("Seleccionar Dispensario", dispensario_options)
else:
    selected_dispensario = "Todos"
    st.sidebar.info("No hay datos de dispensarios disponibles para filtrar")

# Convertir fechas si los dataframes existen
for df in [ventas_df, alertas_df]:
    if df is not None:
        if 'fecha_venta' in df.columns:
            df['fecha_venta'] = pd.to_datetime(df['fecha_venta'], errors='coerce')
        if 'fecha_creacion' in df.columns:
            df['fecha_creacion'] = pd.to_datetime(df['fecha_creacion'], errors='coerce')

# Layout principal
tab1, tab2, tab3, tab4, tab5, tab6 = st.tabs(["Resumen", "Inventario", "Ventas", "Calidad y Alertas", "Vistas Personalizadas", "Editor de Datos"])

with tab1:
    st.header("Resumen General")
    
    # KPIs
    col1, col2, col3, col4 = st.columns(4)
    
    with col1:
        if ventas_df is not None and 'total' in ventas_df.columns:
            try:
                # Asegurarse de que la columna 'total' es numérica
                ventas_df['total'] = pd.to_numeric(ventas_df['total'], errors='coerce')
                total_ventas = ventas_df['total'].sum()
                st.metric("Ventas Totales", f"${total_ventas:,.2f}")
            except Exception as e:
                st.error(f"Error calculando ventas totales: {e}")
                st.metric("Ventas Totales", "Error")
        else:
            st.metric("Ventas Totales", "N/D")
    
    with col2:
        if alertas_df is not None and 'estado' in alertas_df.columns:
            try:
                alertas_activas = alertas_df[alertas_df['estado'] == 'Activa'].shape[0]
                st.metric("Alertas Activas", alertas_activas)
            except Exception as e:
                st.error(f"Error calculando alertas activas: {e}")
                st.metric("Alertas Activas", "Error")
        else:
            st.metric("Alertas Activas", "N/D")
    
    with col3:
        if (inventario_dispensario_df is not None and 
            'cantidad' in inventario_dispensario_df.columns and 
            'stock_minimo' in inventario_dispensario_df.columns):
            try:
                # Asegurarse de que las columnas son numéricas
                inventario_dispensario_df['cantidad'] = pd.to_numeric(inventario_dispensario_df['cantidad'], errors='coerce')
                inventario_dispensario_df['stock_minimo'] = pd.to_numeric(inventario_dispensario_df['stock_minimo'], errors='coerce')
                
                productos_stock_bajo = inventario_dispensario_df[
                    inventario_dispensario_df['cantidad'] <= inventario_dispensario_df['stock_minimo']
                ].shape[0]
                st.metric("Productos con Stock Bajo", productos_stock_bajo)
            except Exception as e:
                st.error(f"Error calculando stock bajo: {e}")
                st.metric("Productos con Stock Bajo", "Error")
        else:
            st.metric("Productos con Stock Bajo", "N/D")
    
    with col4:
        if ventas_df is not None and 'fecha_venta' in ventas_df.columns:
            try:
                # Asegurarse de que la fecha está en formato datetime
                ventas_df['fecha_venta'] = pd.to_datetime(ventas_df['fecha_venta'], errors='coerce')
                ventas_hoy = ventas_df[ventas_df['fecha_venta'].dt.date == datetime.today().date()].shape[0]
                st.metric("Ventas Hoy", ventas_hoy)
            except Exception as e:
                st.error(f"Error calculando ventas de hoy: {e}")
                st.metric("Ventas Hoy", "Error")
        else:
            st.metric("Ventas Hoy", "N/D")
    
    # Gráfico de ventas por día
    st.subheader("Ventas por Día")
    if ventas_df is not None and 'fecha_venta' in ventas_df.columns and 'total' in ventas_df.columns:
        try:
            # Asegurarse de que las columnas tienen el formato correcto
            ventas_df['fecha_venta'] = pd.to_datetime(ventas_df['fecha_venta'], errors='coerce')
            ventas_df['total'] = pd.to_numeric(ventas_df['total'], errors='coerce')
            
            # Eliminar filas con valores NaN
            ventas_df_clean = ventas_df.dropna(subset=['fecha_venta', 'total'])
            
            ventas_por_dia = ventas_df_clean.groupby(ventas_df_clean['fecha_venta'].dt.date)['total'].sum().reset_index()
            fig_ventas = px.line(ventas_por_dia, x='fecha_venta', y='total', 
                                 title='Evolución de Ventas Diarias',
                                 color_discrete_sequence=[color_principal])
            st.plotly_chart(fig_ventas, width='stretch')
        except Exception as e:
            st.error(f"Error generando gráfico de ventas: {e}")
    else:
        st.warning("No hay datos de ventas disponibles para mostrar")
    
    # Top productos
    st.subheader("Productos Más Vendidos")
    if (detalle_venta_df is not None and productos_df is not None and
        'producto_id' in detalle_venta_df.columns and 'id' in productos_df.columns and
        'cantidad' in detalle_venta_df.columns and 'nombre' in productos_df.columns):
        try:
            # Asegurarse de que las columnas son numéricas donde es necesario
            detalle_venta_df['cantidad'] = pd.to_numeric(detalle_venta_df['cantidad'], errors='coerce')
            detalle_venta_df['producto_id'] = pd.to_numeric(detalle_venta_df['producto_id'], errors='coerce')
            productos_df['id'] = pd.to_numeric(productos_df['id'], errors='coerce')
            
            # Eliminar filas con valores NaN
            detalle_venta_df_clean = detalle_venta_df.dropna(subset=['producto_id', 'cantidad'])
            productos_df_clean = productos_df.dropna(subset=['id'])
            
            ventas_productos = pd.merge(detalle_venta_df_clean, productos_df_clean, left_on='producto_id', right_on='id')
            top_productos = ventas_productos.groupby('nombre')['cantidad'].sum().nlargest(5).reset_index()
            fig_productos = px.bar(top_productos, x='nombre', y='cantidad', 
                                   title='Top 5 Productos por Cantidad Vendida',
                                   color_discrete_sequence=[color_secundario])
            st.plotly_chart(fig_productos, width='stretch')
        except Exception as e:
            st.error(f"Error generando gráfico de productos: {e}")
    else:
        st.warning("No hay datos de productos disponibles para mostrar")

with tab2:
    st.header("Gestión de Inventario")
    
    col1, col2 = st.columns(2)
    
    with col1:
        st.subheader("Inventario en Depósito")
        if inventario_deposito_df is not None and productos_df is not None:
            try:
                inventario_deposito = pd.merge(inventario_deposito_df, productos_df, 
                                              left_on='producto_id', right_on='id')
                inventario_deposito['nivel_stock'] = inventario_deposito['cantidad'] / inventario_deposito['stock_maximo'] * 100
                
                fig_deposito = px.bar(inventario_deposito, x='nombre', y='cantidad',
                                     title='Cantidad en Depósito por Producto',
                                     color='nivel_stock',
                                     color_continuous_scale='RdYlGn')
                st.plotly_chart(fig_deposito, width='stretch')
            except Exception as e:
                st.error(f"Error generando gráfico de inventario: {e}")
        else:
            st.warning("No hay datos de inventario en depósito disponibles")
    
    with col2:
        st.subheader("Inventario por Dispensario")
        if inventario_dispensario_df is not None and productos_df is not None and dispensarios_df is not None:
            try:
                # Filtrar por dispensario si se seleccionó uno
                if selected_dispensario != "Todos":
                    dispensario_id = dispensarios_df[dispensarios_df['nombre'] == selected_dispensario]['id'].iloc[0]
                    inventario_filtrado = inventario_dispensario_df[inventario_dispensario_df['dispensario_id'] == dispensario_id]
                else:
                    inventario_filtrado = inventario_dispensario_df
                
                # Unir con datos de productos и dispensarios
                inventario_filtrado = pd.merge(inventario_filtrado, productos_df, 
                                              left_on='producto_id', right_on='id')
                inventario_filtrado = pd.merge(inventario_filtrado, dispensarios_df, 
                                              left_on='dispensario_id', right_on='id', 
                                              suffixes=('_producto', '_dispensario'))
                
                fig_dispensario = px.bar(inventario_filtrado, x='nombre_producto', y='cantidad',
                                        color='nombre_dispensario',
                                        title=f'Inventario por Producto ({selected_dispensario})')
                st.plotly_chart(fig_dispensario, width='stretch')
            except Exception as e:
                st.error(f"Error generando gráfico de inventario por dispensario: {e}")
        else:
            st.warning("No hay datos de inventario por dispensario disponibles")
    
    # Productos con stock crítico
    st.subheader("Productos con Stock Crítico")
    if inventario_dispensario_df is not None and productos_df is not None and dispensarios_df is not None:
        try:
            stock_critico = inventario_dispensario_df[
                inventario_dispensario_df['cantidad'] <= inventario_dispensario_df['stock_minimo']
            ]
            
            # Realizar los merges correctamente
            stock_critico = pd.merge(stock_critico, productos_df, left_on='producto_id', right_on='id')
            stock_critico = pd.merge(stock_critico, dispensarios_df, left_on='dispensario_id', right_on='id')
            
            # Renombrar columnas para mayor claridad
            stock_critico = stock_critico.rename(columns={
                'nombre_x': 'nombre_producto',
                'nombre_y': 'nombre_dispensario'
            })
            
            if not stock_critico.empty:
                st.dataframe(stock_critico[['nombre_dispensario', 'nombre_producto', 'cantidad', 'stock_minimo']], width='stretch')
            else:
                st.success("No hay productos con stock crítico")
        except Exception as e:
            st.error(f"Error cargando datos de stock crítico: {e}")
    else:
        st.warning("No hay datos disponibles para mostrar stock crítico")

with tab3:
    st.header("Análisis de Ventas")
    
    col1, col2 = st.columns(2)
    
    with col1:
        st.subheader("Ventas por Método de Pago")
        if ventas_df is not None and 'metodo_pago' in ventas_df.columns and 'total' in ventas_df.columns:
            try:
                ventas_metodo = ventas_df.groupby('metodo_pago')['total'].sum().reset_index()
                fig_metodo = px.pie(ventas_metodo, values='total', names='metodo_pago',
                                   title='Distribución de Ventas por Método de Pago')
                st.plotly_chart(fig_metodo, width='stretch')
            except Exception as e:
                st.error(f"Error generando gráfico de métodos de pago: {e}")
        else:
            st.warning("No hay datos de métodos de pago disponibles")
    
    with col2:
        st.subheader("Ventas por Dispensario")
        if ventas_df is not None and dispensarios_df is not None and 'dispensario_id' in ventas_df.columns:
            try:
                ventas_dispensario = pd.merge(ventas_df, dispensarios_df, left_on='dispensario_id', right_on='id')
                ventas_por_dispensario = ventas_dispensario.groupby('nombre')['total'].sum().reset_index()
                fig_dispensario_ventas = px.bar(ventas_por_dispensario, x='nombre', y='total',
                                               title='Ventas Totales por Dispensario',
                                               color_discrete_sequence=[color_principal])
                st.plotly_chart(fig_dispensario_ventas, width='stretch')
            except Exception as e:
                st.error(f"Error generando gráfico de ventas por dispensario: {e}")
        else:
            st.warning("No hay datos de ventas por dispensario disponibles")
    
    # Top clientes
    st.subheader("Top 5 Clientes por Consumo")
    if ventas_df is not None and clientes_df is not None and 'cliente_id' in ventas_df.columns:
        try:
            ventas_clientes = pd.merge(ventas_df, clientes_df, left_on='cliente_id', right_on='id')
            top_clientes = ventas_clientes.groupby(['nombre', 'apellido'])['total'].sum().nlargest(5).reset_index()
            top_clientes['nombre_completo'] = top_clientes['nombre'] + ' ' + top_clientes['apellido']
            fig_clientes = px.bar(top_clientes, x='nombre_completo', y='total',
                                 title='Top 5 Clientes por Monto Gastado',
                                 color_discrete_sequence=[color_secundario])
            st.plotly_chart(fig_clientes, width='stretch')
        except Exception as e:
            st.error(f"Error generando gráfico de clientes: {e}")
    else:
        st.warning("No hay datos de clientes disponibles")

with tab4:
    st.header("Control de Calidad y Alertas")
    
    col1, col2 = st.columns(2)
    
    with col1:
        st.subheader("Estado de Alertas")
        if alertas_df is not None and 'estado' in alertas_df.columns:
            try:
                estado_alertas = alertas_df.groupby('estado').size().reset_index(name='count')
                fig_alertas = px.pie(estado_alertas, values='count', names='estado',
                                    title='Distribución de Alertas por Estado')
                st.plotly_chart(fig_alertas, width='stretch')
            except Exception as e:
                st.error(f"Error generando gráfico de alertas: {e}")
        else:
            st.warning("No hay datos de alertas disponibles")
        
        # Alertas activas por prioridad
        if alertas_df is not None and 'estado' in alertas_df.columns and 'prioridad' in alertas_df.columns:
            try:
                alertas_activas = alertas_df[alertas_df['estado'] == 'Activa']
                if not alertas_activas.empty:
                    st.subheader("Alertas Activas por Prioridad")
                    alertas_prioridad = alertas_activas.groupby('prioridad').size().reset_index(name='count')
                    fig_prioridad = px.bar(alertas_prioridad, x='prioridad', y='count',
                                          title='Alertas Activas por Nivel de Prioridad',
                                          color_discrete_sequence=[color_principal])
                    st.plotly_chart(fig_prioridad, width='stretch')
            except Exception as e:
                st.error(f"Error generando gráfico de prioridad de alertas: {e}")
        else:
            st.warning("No hay datos de alertas disponibles")
    
    with col2:
        st.subheader("Resultados de Control de Calidad")
        if control_calidad_df is not None and 'resultado' in control_calidad_df.columns:
            try:
                resultados_calidad = control_calidad_df.groupby('resultado').size().reset_index(name='count')
                fig_calidad = px.pie(resultados_calidad, values='count', names='resultado',
                                    title='Distribución de Resultados de Control de Calidad')
                st.plotly_chart(fig_calidad, width='stretch')
            except Exception as e:
                st.error(f"Error generando gráfico de control de calidad: {e}")
        else:
            st.warning("No hay datos de control de calidad disponibles")
    
    # Detalle de alertas activas
    st.subheader("Detalle de Alertas Activas")
    if alertas_df is not None and productos_df is not None:
        try:
            alertas_detalle = alertas_df[alertas_df['estado'] == 'Activa']
            if not alertas_detalle.empty:
                # Unir solo con las columnas necesarias de productos
                alertas_detalle = pd.merge(alertas_detalle, productos_df[['id', 'nombre']], 
                                          left_on='producto_id', right_on='id', how='left')
                
                # Seleccionar solo las columnas que existen
                columnas_disponibles = []
                for col in ['tipo_alerta', 'mensaje', 'prioridad', 'nombre', 'fecha_creacion']:
                    if col in alertas_detalle.columns:
                        columnas_disponibles.append(col)
                
                if columnas_disponibles:
                    st.dataframe(alertas_detalle[columnas_disponibles], width='stretch')
                else:
                    st.warning("No hay columnas disponibles para mostrar")
            else:
                st.success("No hay alertas activas en este momento")
        except Exception as e:
            st.error(f"Error al cargar alertas: {e}")
    else:
        st.warning("No hay datos de alertas o productos disponibles")

with tab5:
    st.header("Vistas Personalizadas")
    
    # Mostrar todas las hojas disponibles
    st.subheader("Hojas de datos disponibles")
    for nombre_hoja, df in excel_data.items():
        with st.expander(f"Hoja: {nombre_hoja}"):
            st.write(f"Filas: {df.shape[0]}, Columnas: {df.shape[1]}")
            st.dataframe(df.head(), width='stretch')
    
    # Crear vistas personalizadas
    st.subheader("Crear vista personalizada")
    
    col1, col2 = st.columns(2)
    
    with col1:
        hoja_seleccionada = st.selectbox(
            "Seleccionar hoja de datos",
            options=list(excel_data.keys())
        )
    
    with col2:
        if hoja_seleccionada:
            df_seleccionado = excel_data[hoja_seleccionada]
            columnas_seleccionadas = st.multiselect(
                "Seleccionar columnas",
                options=list(df_seleccionado.columns),
                default=list(df_seleccionado.columns)
            )
    
    if hoja_seleccionada and columnas_seleccionadas:
        df_filtrado = excel_data[hoja_seleccionada][columnas_seleccionadas]
        st.dataframe(df_filtrado, width='stretch')
        
        # Opciones de visualización
        tipo_grafico = st.selectbox(
            "Tipo de gráfico",
            options=["Ninguno", "Barras", "Líneas", "Pastel", "Dispersión"]
        )
        
        if tipo_grafico != "Ninguno":
            col_x = st.selectbox("Columna para eje X", options=columnas_seleccionadas)
            col_y = st.selectbox("Columna para eje Y", options=columnas_seleccionadas)
            
            if st.button("Generar gráfico"):
                try:
                    if tipo_grafico == "Barras":
                        fig = px.bar(df_filtrado, x=col_x, y=col_y, title=f"{col_y} por {col_x}")
                    elif tipo_grafico == "Líneas":
                        fig = px.line(df_filtrado, x=col_x, y=col_y, title=f"{col_y} por {col_x}")
                    elif tipo_grafico == "Pastel":
                        fig = px.pie(df_filtrado, names=col_x, values=col_y, title=f"Distribución de {col_y} por {col_x}")
                    elif tipo_grafico == "Dispersión":
                        fig = px.scatter(df_filtrado, x=col_x, y=col_y, title=f"{col_y} vs {col_x}")
                    
                    st.plotly_chart(fig, width='stretch')
                except Exception as e:
                    st.error(f"Error al generar gráfico: {e}")

with tab6:
    st.header("Editor de Datos")
    st.markdown("Esta sección permite editar directamente los datos del archivo Excel.")
    
    # Seleccionar archivo para editar
    archivos_opciones = ["data/db.xlsx"]
    if archivos_cargados:
        archivos_opciones.extend([archivo.name for archivo in archivos_cargados])
    
    archivo_editar = st.selectbox(
        "Seleccionar archivo para editar",
        options=archivos_opciones
    )
    
    # Cargar el archivo seleccionado
    archivo_data = {}
    try:
        if archivo_editar == "data/db.xlsx":
            if os.path.exists('data/db.xlsx'):
                archivo_data = pd.read_excel('data/db.xlsx', sheet_name=None)
            else:
                st.warning("El archivo data/db.xlsx no existe. Se creará uno nuevo al guardar.")
        else:
            # Buscar el archivo cargado
            for archivo in archivos_cargados:
                if archivo.name == archivo_editar:
                    archivo_data = pd.read_excel(archivo, sheet_name=None)
                    break
    except Exception as e:
        st.error(f"Error cargando {archivo_editar}: {e}")
    
    # Seleccionar hoja para editar
    if archivo_data:
        hojas_disponibles = list(archivo_data.keys())
        hoja_seleccionada = st.selectbox("Seleccionar hoja para editar", hojas_disponibles)
        
        # Mostrar datos actuales
        st.subheader(f"Datos actuales en {hoja_seleccionada}")
        df_actual = archivo_data[hoja_seleccionada]
        edited_df = st.data_editor(df_actual, num_rows="dynamic", width='stretch')
        
        # Botones para guardar cambios
        col1, col2, col3 = st.columns(3)
        
        with col1:
            if st.button("💾 Guardar cambios", width='stretch'):
                if save_to_excel(edited_df, hoja_seleccionada, archivo_editar):
                    # Limpiar caché para recargar datos
                    st.rerun()
        
        with col2:
            if st.button("🔄 Restaurar original", width='stretch'):
                st.rerun()
        
        with col3:
            if st.button("📥 Descargar como Excel", width='stretch'):
                buffer = io.BytesIO()
                with pd.ExcelWriter(buffer, engine='xlsxwriter') as writer:
                    edited_df.to_excel(writer, sheet_name=hoja_seleccionada, index=False)
                st.download_button(
                    label="Descargar",
                    data=buffer.getvalue(),
                    file_name=f"{hoja_seleccionada}_editado.xlsx",
                    mime="application/vnd.ms-excel",
                    width='stretch'
                )
        
        # Sección para agregar nuevas columnas
        st.subheader("Agregar nueva columna")
        col_nombre = st.text_input("Nombre de la nueva columna")
        col_tipo = st.selectbox("Tipo de datos", ["Texto", "Número", "Fecha", "Booleano"])
        
        if st.button("➕ Agregar columna"):
            if col_nombre:
                if col_tipo == "Texto":
                    edited_df[col_nombre] = ""
                elif col_tipo == "Número":
                    edited_df[col_nombre] = 0
                elif col_tipo == "Fecha":
                    edited_df[col_nombre] = pd.Timestamp.now()
                elif col_tipo == "Booleano":
                    edited_df[col_nombre] = False
                
                st.success(f"Columna '{col_nombre}' agregada. Recuerde guardar los cambios.")
            else:
                st.warning("Por favor, ingrese un nombre para la columna")
        
        # Sección para crear nuevas hojas
        st.subheader("Crear nueva hoja")
        nueva_hoja_nombre = st.text_input("Nombre de la nueva hoja")
        
        if st.button("📄 Crear nueva hoja"):
            if nueva_hoja_nombre:
                # Crear un DataFrame vacío con una columna por defecto
                nuevo_df = pd.DataFrame(columns=['ID'])
                if save_to_excel(nuevo_df, nueva_hoja_nombre, archivo_editar):
                    st.rerun()
            else:
                st.warning("Por favor, ingrese un nombre para la nueva hoja")
    
    else:
        st.warning("No hay datos disponibles para editar.")

# Footer
st.markdown("---")
st.markdown("© 2024 Sistema de Gestión de Cannabis Medicinal - Todos los derechos reservados")